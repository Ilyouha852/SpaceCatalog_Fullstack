from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from django.utils import timezone
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.response import Response
from django.conf import settings
import os


class ExcelExportMixin:

    @action(detail=False, methods=['get'], url_path='export_excel', url_name='export_excel')
    def export_excel(self, request):
        queryset = self.get_queryset()
        serializer = self.get_serializer(queryset, many=True)
        data = serializer.data

        wb = Workbook()
        ws = wb.active
        model_name = self.queryset.model._meta.verbose_name_plural.title()
        ws.title = model_name[:31]
        if data:
            headers = list(data[0].keys())
        else:
            headers = [field.name for field in self.queryset.model._meta.fields]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1e40af")
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header.title().replace('_', ' '))
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_num)].width = 22

        for row_num, item in enumerate(data, 2):
            for col_num, key in enumerate(headers, 1):
                value = item.get(key, "")
                if isinstance(value, dict) and 'name' in value:
                    value = value['name']
                elif isinstance(value, dict) and 'first_name' in value:
                    value = f"{value['first_name']} {value['last_name']}"
                ws.cell(row=row_num, column=col_num, value=str(value or ""))

        filename = f"{model_name}_{timezone.now():%Y%m%d_%H%M%S}.xlsx"
        
        downloads_path = os.path.join(os.path.expanduser("~"), "Downloads")
        if not os.path.exists(downloads_path):
            os.makedirs(downloads_path)
        file_path = os.path.join(downloads_path, filename)
        wb.save(file_path)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response