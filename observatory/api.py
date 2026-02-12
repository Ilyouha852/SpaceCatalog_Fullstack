from rest_framework.viewsets import GenericViewSet
from rest_framework import mixins
from rest_framework.decorators import action
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from rest_framework.response import Response
from rest_framework import serializers
from rest_framework.permissions import IsAuthenticated
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import io
from .models import Observatory, Astronomer, Researcher, Observation, SpaceObject
from .serializers import (
    ObservatorySerializer,
    AstronomerSerializer,
    ResearcherSerializer,
    ObservationSerializer,
    SpaceObjectSerializer,
)
from .permissions import (
    IsAstronomerPermission,
    IsResearcherPermission,
    IsAdminPermission,
    CanSeeStatisticsPermission,
    CanManageAstronomersPermission,
    CanManageResearchersPermission,
    SecondFactorPermission,
    CanCreateObservationsPermission,
)
from django.db.models import Count

class ObservatoryViewSet(
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    mixins.ListModelMixin,
    GenericViewSet
):
    queryset = Observatory.objects.all()
    serializer_class = ObservatorySerializer
    
    def get_permissions(self):
        if self.action in ['update', 'destroy']:
             return [SecondFactorPermission()]
        elif self.action in ['create']:
            return [IsAdminPermission()]
        return super().get_permissions()

    @action(detail=False, methods=["GET"], url_path="stats", permission_classes=[CanSeeStatisticsPermission])
    def get_stats(self, request, *args, **kwargs):
        aggregate_stats = {
            "total_count": Observatory.objects.count(),
            "astronomers_count": Astronomer.objects.count(),
        }

        observatory_stats = Observatory.objects.annotate(
            astronomer_count=Count('astronomer')
        ).order_by('-astronomer_count').values('id', 'name', 'astronomer_count')[:5]
        
        return Response({
            "aggregate_stats": aggregate_stats,
            "observatory_stats": list(observatory_stats)
        })
    @action(detail=False, methods=["GET"], url_path="export-excel")
    def export_excel(self, request, *args, **kwargs):
        """
        Экспорт данных обсерваторий в Excel
        """

        observatories = Observatory.objects.all()

        
        wb = Workbook()
        ws = wb.active
        ws.title = "Обсерватории"

        headers = ["ID", "Название", "Телефон", "Адрес"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
            ws[f"{col_letter}1"].font = ws[f"{col_letter}1"].font.copy(bold=True)

        for row_num, observatory in enumerate(observatories, 2):
            ws[f"A{row_num}"] = observatory.id
            ws[f"B{row_num}"] = observatory.name
            ws[f"C{row_num}"] = observatory.phone if observatory.phone else ""
            ws[f"D{row_num}"] = observatory.address if hasattr(observatory, 'address') and observatory.address else ""

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
 
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="observatories.xlsx"'
        
        return response
class AstronomerViewSet(
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    mixins.ListModelMixin,
    GenericViewSet
):
    queryset = Astronomer.objects.all()
    serializer_class = AstronomerSerializer
    
    def get_permissions(self):
        if self.action in ['update', 'destroy']:
            return [SecondFactorPermission()]
        elif self.action in ['create']:
            return [CanManageAstronomersPermission()]
        return super().get_permissions()

    @action(detail=False, methods=["GET"], url_path="stats", permission_classes=[CanSeeStatisticsPermission])
    def get_stats(self, request, *args, **kwargs):
        aggregate_stats = {
            "total_count": Astronomer.objects.count(),
            "research_fields_count": Astronomer.objects.values('research_field').distinct().count(),
            "observatories_count": Astronomer.objects.values('observatory').distinct().count(),
            "observations_count": Observation.objects.count(),
        }

        specialization_stats = Astronomer.objects.values('research_field').annotate(
            count=Count('id')
        ).order_by('-count')
        
        return Response({
            "aggregate_stats": aggregate_stats,
            "specialization_stats": list(specialization_stats)
        })
    @action(detail=False, methods=["GET"], url_path="export-excel")
    def export_excel(self, request, *args, **kwargs):
        """
        Экспорт данных астрономов в Excel
        """
        astronomers = Astronomer.objects.select_related('observatory').all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Астрономы"

        headers = ["ID", "ФИО", "Поле исследований", "Обсерватория"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
            ws[f"{col_letter}1"].font = ws[f"{col_letter}1"].font.copy(bold=True)

        for row_num, a in enumerate(astronomers, 2):
            ws[f"A{row_num}"] = a.id
            ws[f"B{row_num}"] = a.name
            ws[f"C{row_num}"] = a.research_field or ""
            ws[f"D{row_num}"] = a.observatory.name if a.observatory else ""

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="astronomers.xlsx"'
        return response
    
    @action(detail=False, url_path="my-appointments", methods=['GET'], permission_classes=[IsAstronomerPermission])
    def my_observations(self, request, *args, **kwargs):
        astronomer = request.user.astronomer
        appointments = Observation.objects.filter(astronomer=astronomer)
        serializer = ObservationSerializer(appointments, many=True)
        return Response(serializer.data)
    
    @action(detail=False, url_path="my-researchers", methods=['GET'], permission_classes=[IsAstronomerPermission])
    def my_researchers(self, request, *args, **kwargs):
        astronomer = request.user.astronomer
        researcher_ids = Observation.objects.filter(astronomer=astronomer).values_list('researcher', flat=True).distinct()
        researchers = Researcher.objects.filter(id__in=researcher_ids)
        serializer = ResearcherSerializer(researchers, many=True)
        return Response(serializer.data)
    
    @action(detail=False, url_path="my-observation-records", methods=['GET'], permission_classes=[IsAstronomerPermission])
    def my_observation_records(self, request, *args, **kwargs):
        astronomer = request.user.astronomer
        objects = SpaceObject.objects.filter(astronomer=astronomer)
        serializer = SpaceObjectSerializer(objects, many=True)
        return Response(serializer.data)

class ResearcherViewSet(
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    mixins.ListModelMixin,
    GenericViewSet
):
    queryset = Researcher.objects.all()
    serializer_class = ResearcherSerializer
    
    def get_permissions(self):
        if self.action in ['update', 'destroy']:
            return [SecondFactorPermission()]
        elif self.action in ['create']:
            return [CanManageResearchersPermission()]
        elif self.action == 'list':
             return [IsAuthenticated()]
        return super().get_permissions()

    @action(detail=False, methods=["GET"], url_path="stats", permission_classes=[CanSeeStatisticsPermission])
    def get_stats(self, request, *args, **kwargs):
        aggregate_stats = {
            "total_count": Researcher.objects.count(),
            "with_user_account": Researcher.objects.filter(user__isnull=False).count(),
            "observations_count": Observation.objects.filter(researcher__isnull=False).count(),
        }

        top_patients = Researcher.objects.annotate(
            observation_count=Count('observation')
        ).order_by('-observation_count').values('id', 'name', 'observation_count')[:5]
        
        return Response({
            "aggregate_stats": aggregate_stats,
            "top_patients_by_appointments": list(top_patients)
        })
    
    @action(detail=False, url_path="my-appointments", methods=['GET'], permission_classes=[IsResearcherPermission])
    def my_observations(self, request, *args, **kwargs):
        researcher = request.user.researcher
        appointments = Observation.objects.filter(researcher=researcher)
        serializer = ObservationSerializer(appointments, many=True)
        return Response(serializer.data)
    
    @action(detail=False, url_path="my-medical-records", methods=['GET'], permission_classes=[IsResearcherPermission])
    def my_observation_records(self, request, *args, **kwargs):
        # Some users may not have a related Researcher object — handle gracefully
        if not hasattr(request.user, 'researcher'):
            return Response([])
        # Researcher doesn't own SpaceObjects; return empty list for compatibility
        return Response([])
    @action(detail=False, methods=["GET"], url_path="export-excel")
    def export_excel(self, request, *args, **kwargs):
        """
        Экспорт данных исследователей в Excel
        """
        researchers = Researcher.objects.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Исследователи"

        headers = ["ID", "ФИО", "Дата рождения", "Телефон", "User ID"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
            ws[f"{col_letter}1"].font = ws[f"{col_letter}1"].font.copy(bold=True)

        for row_num, r in enumerate(researchers, 2):
            ws[f"A{row_num}"] = r.id
            ws[f"B{row_num}"] = r.name
            ws[f"C{row_num}"] = r.birth_date.isoformat() if r.birth_date else ""
            ws[f"D{row_num}"] = r.phone or ""
            ws[f"E{row_num}"] = r.user.id if r.user_id else ""

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="researchers.xlsx"'
        return response

class ObservationViewSet(
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    mixins.ListModelMixin,
    GenericViewSet
):
    queryset = Observation.objects.all()
    serializer_class = ObservationSerializer
    
    def get_permissions(self):
        if self.action in ['update', 'destroy']:
            return [SecondFactorPermission()]
        elif self.action in ['create']:
            return [CanCreateObservationsPermission()]
        return super().get_permissions()

    def get_queryset(self):
        qs = super().get_queryset()

        if self.request.user.is_superuser:
            return qs

        if hasattr(self.request.user, 'astronomer'):
            astronomer = self.request.user.astronomer
            return qs.filter(astronomer=astronomer)
 
        if hasattr(self.request.user, 'researcher'):
            researcher = self.request.user.researcher
            return qs.filter(researcher=researcher)
 
        return qs.none()
    @action(detail=False, methods=["GET"], url_path="stats", permission_classes=[CanSeeStatisticsPermission])
    def get_stats(self, request, *args, **kwargs):
        from django.db.models import Count
        
        aggregate_stats = {
            "total_count": Observation.objects.count(),
            "completed_count": Observation.objects.filter(status='completed').count(),
            "planned_count": Observation.objects.filter(status='planned').count(),
            "cancelled_count": Observation.objects.filter(status='cancelled').count(),
            "pending_count": Observation.objects.filter(status='pending').count(),
        }

        status_stats = Observation.objects.values('status').annotate(
            count=Count('id')
        ).order_by('-count')
 
        top_astronomers = Observation.objects.values(
            'astronomer__id', 'astronomer__name'
        ).annotate(
            appointment_count=Count('id')
        ).order_by('-appointment_count')[:5]

        return Response({
            "aggregate_stats": aggregate_stats,
            "status_stats": list(status_stats),
            "top_astronomers_by_observations": list(top_astronomers)
        })
    @action(detail=False, methods=["GET"], url_path="export-excel")
    def export_excel(self, request, *args, **kwargs):
        """
        Экспорт данных наблюдений в Excel
        """
        observations = Observation.objects.select_related('astronomer', 'researcher').all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Наблюдения"

        headers = ["ID", "Дата и время", "Статус", "Астроном", "Исследователь"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
            ws[f"{col_letter}1"].font = ws[f"{col_letter}1"].font.copy(bold=True)

        for row_num, o in enumerate(observations, 2):
            ws[f"A{row_num}"] = o.id
            ws[f"B{row_num}"] = o.date_time.isoformat() if o.date_time else ""
            ws[f"C{row_num}"] = o.status or ""
            ws[f"D{row_num}"] = o.astronomer.name if getattr(o, 'astronomer', None) else ""
            ws[f"E{row_num}"] = o.researcher.name if getattr(o, 'researcher', None) else ""

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="observations.xlsx"'
        return response

# ObservationRecordViewSet removed — model deleted and data cleared


class SpaceObjectViewSet(
    mixins.CreateModelMixin,
    mixins.RetrieveModelMixin,
    mixins.UpdateModelMixin,
    mixins.DestroyModelMixin,
    mixins.ListModelMixin,
    GenericViewSet
):
    queryset = SpaceObject.objects.all()
    serializer_class = SpaceObjectSerializer

    def get_permissions(self):
        if self.action in ['update', 'destroy']:
            return [SecondFactorPermission()]
        elif self.action in ['create']:
            return [IsAdminPermission()]
        return super().get_permissions()

    def get_queryset(self):
        qs = super().get_queryset()

        if self.request.user.is_superuser:
            return qs

        if hasattr(self.request.user, 'astronomer'):
            astronomer = self.request.user.astronomer
            return qs.filter(astronomer=astronomer)

        return qs.none()

    @action(detail=False, methods=["GET"], url_path="stats", permission_classes=[CanSeeStatisticsPermission])
    def get_stats(self, request, *args, **kwargs):
        aggregate_stats = {
            "total_count": SpaceObject.objects.count(),
            "unique_astronomers": SpaceObject.objects.values('astronomer').distinct().count(),
        }

        top_astronomers = SpaceObject.objects.values(
            'astronomer__id', 'astronomer__name'
        ).annotate(
            object_count=Count('id')
        ).order_by('-object_count')[:5]

        return Response({
            "aggregate_stats": aggregate_stats,
            "top_astronomers_by_objects": list(top_astronomers)
        })
    @action(detail=False, methods=["GET"], url_path="export-excel")
    def export_excel(self, request, *args, **kwargs):
        """
        Экспорт данных космических объектов в Excel
        """
        objects = SpaceObject.objects.select_related('astronomer').all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Космические объекты"

        headers = ["ID", "Имя", "Тип", "Астроном"]
        for col_num, header in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = header
            ws[f"{col_letter}1"].font = ws[f"{col_letter}1"].font.copy(bold=True)

        for row_num, o in enumerate(objects, 2):
            ws[f"A{row_num}"] = o.id
            ws[f"B{row_num}"] = o.name
            ws[f"C{row_num}"] = o.object_type or ""
            ws[f"D{row_num}"] = o.astronomer.name if o.astronomer else ""

        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="space_objects.xlsx"'
        return response
